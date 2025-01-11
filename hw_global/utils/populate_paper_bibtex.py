import os
import requests
import openpyxl
import logging
from datetime import datetime
import json
import re
from doi2bib.crossref import get_bib

# Configure logging
log_filename = f"populate_bibtex_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_filename),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

def extract_doi_from_url(doi_url):
    """
    Extracts the DOI from a URL using regex.
    Args:
        doi_url: URL containing the DOI.
    Returns:
        The DOI string if found, otherwise None.
    """
    if not doi_url:
        return None
        
    try:
        # Regex to find DOI (improved to handle more variations)
        doi_regex = r"(10[.][0-9]{4,}(?:[.][0-9]+)*/(?:(?![\"&\'<>])\S)+)"
        match = re.search(doi_regex, doi_url)
        if match:
            doi = match.group(0)
            logger.info(f"Found DOI: {doi}")
            return doi
        else:
            logger.warning(f"No DOI found in {doi_url}")
            return None
    except Exception as e:
        logger.error(f"Error extracting DOI from {doi_url}: {str(e)}")
        return None

def get_paper_info_from_doi(doi):
    """
    Fetches paper information from a DOI using doi2bib.
    """
    try:
        logger.info(f"Fetching information for DOI: {doi}")
        bib_entry = get_bib(doi)
        if bib_entry and isinstance(bib_entry, tuple):
            status_code, response_text = bib_entry
            logger.info("Received response from doi2bib")
            
            if status_code == True and response_text:
                # Parse the BibTeX response text
                title_match = re.search(r'title\s*=\s*[{"](.+?)[}"],?', response_text)
                author_match = re.search(r'author\s*=\s*[{"](.+?)[}"],?', response_text)
                year_match = re.search(r'year\s*=\s*[{"]?(\d{4})[}"]?,?', response_text)
                journal_match = re.search(r'journal\s*=\s*[{"](.+?)[}"],?', response_text)
                
                if not journal_match:
                    journal_match = re.search(r'journaltitle\s*=\s*[{"](.+?)[}"],?', response_text)
                
                return {
                    "title": title_match.group(1) if title_match else None,
                    "author": author_match.group(1) if author_match else None,
                    "year": year_match.group(1) if year_match else None,
                    "journal": journal_match.group(1) if journal_match else None,
                    "bibtex": response_text
                }
            else:
                logger.warning(f"Invalid status code or empty response text from doi2bib")
                return None
        logger.warning("No valid response from doi2bib")
        return None
    except Exception as e:
        logger.error(f"Error fetching info for DOI {doi}: {str(e)}")
        return None

def remove_doi_url_from_bibtex(bibtex):
    """
    Removes DOI and URL fields from BibTeX entry.
    """
    if not bibtex or bibtex == "Not Found":
        return bibtex
        
    try:
        # Remove url={...}, and one space after it
        bibtex = re.sub(r'url={[^}]+},\s', '', bibtex)
        
        # Remove DOI={...}, and one space after it
        bibtex = re.sub(r'DOI={[^}]+},\s', '', bibtex)
        
        return bibtex.strip()
    except Exception as e:
        logger.error(f"Error removing DOI/URL from BibTeX: {str(e)}")
        return bibtex

def process_excel_file(input_file="paper_information.xlsx", output_file="paper_information_out.xlsx"):
    """
    Processes the input Excel file and creates an output file with DOI responses.
    """
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(input_file)
        sheet = wb.active
        
        # Create new workbook for output
        new_wb = openpyxl.Workbook()
        new_sheet = new_wb.active
        
        # Get headers and ensure Filename and DOI columns exist
        headers = [cell.value for cell in sheet[1]]
        try:
            filename_idx = headers.index('Filename')
            doi_idx = headers.index('DOI')
        except ValueError as e:
            logger.error("Required column 'Filename' or 'DOI' not found in input file")
            raise
            
        # Create new headers with only Filename, DOI and the new fields
        new_headers = ['Filename', 'DOI', 'title', 'author', 'year', 'journal', 'bibtex', 'bibtex_nodoi', 'raw_response']
        new_sheet.append(new_headers)
        
        # Process each row
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            row_data = [cell.value for cell in row]
            filename = row_data[filename_idx]
            doi_url = row_data[doi_idx]
            
            logger.info(f"Processing file: {filename}")
            
            # Initialize empty values for new columns
            title = author = year = journal = bibtex = bibtex_nodoi = raw_response = "Not Found"
            
            if doi_url:
                doi = extract_doi_from_url(doi_url)
                if doi:
                    logger.info(f"Processing DOI: {doi}")
                    response_data = get_paper_info_from_doi(doi)
                    
                    if response_data:
                        # Extract individual fields
                        title = response_data.get('title', 'Not Found')
                        author = response_data.get('author', 'Not Found')
                        year = response_data.get('year', 'Not Found')
                        journal = response_data.get('journal', 'Not Found')
                        bibtex = response_data.get('bibtex', 'Not Found')
                        bibtex_nodoi = remove_doi_url_from_bibtex(bibtex)
                        raw_response = json.dumps(response_data, ensure_ascii=False)
                    else:
                        raw_response = "Failed to fetch data"
                else:
                    raw_response = "Failed to extract DOI from URL"
            else:
                raw_response = "No DOI provided"
                
            # Only include Filename, DOI and the new fields
            new_row = [filename, doi_url, title, author, year, journal, bibtex, bibtex_nodoi, raw_response]
            new_sheet.append(new_row)
            
            # Log progress
            if row_idx % 10 == 0:
                logger.info(f"Processed {row_idx} rows")
        
        # Adjust column widths
        for column in new_sheet.columns:
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 100)  # Cap width at 100 characters
            new_sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save the new workbook
        new_wb.save(output_file)
        logger.info(f"Successfully created {output_file}")
        
    except Exception as e:
        logger.error(f"Error processing Excel file: {str(e)}")
        raise

def main():
    """
    Main function to run the script.
    """
    logger.info("Starting paper bibtex population process")
    try:
        process_excel_file()
        logger.info("Process completed successfully")
    except Exception as e:
        logger.error(f"Process failed: {str(e)}")

if __name__ == "__main__":
    main()
